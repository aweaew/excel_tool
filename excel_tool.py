import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import openpyxl
from openpyxl.utils import get_column_letter
import os
import shutil
import datetime
import threading
import glob
from collections import Counter

# 版本号
APP_VERSION = "V40 (终极融合·全功能完整版)"

# 尝试导入 win32com
try:
    import win32com.client as win32
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

class ExcelToolApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"各区表格协同工具 {APP_VERSION}")
        self.root.geometry("1150x950")

        # 样式设置
        self.style = ttk.Style()
        try: self.style.theme_use('clam')
        except: pass
        self.style.configure("Treeview", rowheight=28, font=("Microsoft YaHei", 9))
        self.style.configure("Treeview.Heading", font=("Microsoft YaHei", 9, "bold"), background="#d9d9d9")

        # 核心缓存
        self.preview_cache = {"max_row": 0, "max_col": 0}
        self.merge_files_cache = [] 
        self.current_template = None 
        self.file_stats_cache = {} # V39: 物理行数缓存
        
        # 引擎状态
        self.has_excel = False; self.has_wps = False; self.engine_choice = tk.StringVar(value="auto")

        # --- 1. 顶部引擎 ---
        self.init_engine_panel()

        # --- 2. 选项卡 ---
        self.tab_control = ttk.Notebook(root)
        self.tab_split = ttk.Frame(self.tab_control)
        self.tab_merge = ttk.Frame(self.tab_control)
        self.tab_control.add(self.tab_split, text=' 第一步：拆分与工具 ')
        self.tab_control.add(self.tab_merge, text=' 第二步：合并回收 ')
        self.tab_control.pack(expand=0, fill="x", padx=5, pady=(5, 0))

        self.init_split_tab()
        self.init_merge_tab()

        # --- 3. 公共预览区 (V40: 找回控制条) ---
        frame_preview = tk.LabelFrame(root, text="表格预览", padx=10, pady=5)
        frame_preview.pack(fill="both", expand=True, padx=10, pady=(0, 5))
        
        # [V40找回] 预览控制条
        frame_prev_ctrl = tk.Frame(frame_preview)
        frame_prev_ctrl.pack(fill="x", pady=(0, 5))
        
        tk.Label(frame_prev_ctrl, text="预览行数限制:", font=("Microsoft YaHei", 9)).pack(side="left")
        self.preview_limit = tk.IntVar(value=50) # 默认50行
        self.spin_limit = tk.Spinbox(frame_prev_ctrl, from_=10, to=10000, textvariable=self.preview_limit, width=6)
        self.spin_limit.pack(side="left", padx=5)
        
        tk.Button(frame_prev_ctrl, text="⟳ 刷新预览", command=self.refresh_preview, height=1, bg="#E0FFFF").pack(side="left", padx=5)
        
        self.lbl_preview_info = tk.Label(frame_prev_ctrl, text="[暂无预览]", fg="gray", anchor="w")
        self.lbl_preview_info.pack(side="left", padx=10)

        # 预览表格
        self.tree = ttk.Treeview(frame_preview, show='headings', height=8)
        self.tree.tag_configure('odd', background='white')
        self.tree.tag_configure('even', background='#f2f5f9')
        
        vsb = ttk.Scrollbar(frame_preview, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frame_preview, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        
        self.tree.bind('<ButtonRelease-1>', self.on_preview_click)
        self.current_preview_file = None # 记录当前预览文件以便刷新

        # --- 4. 日志 ---
        log_frame = tk.Frame(root)
        log_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(log_frame, text=f"版本: {APP_VERSION}", fg="gray", anchor="e").pack(side="right", padx=10)
        tk.Label(log_frame, text="执行日志:", font=("Microsoft YaHei", 9, "bold"), anchor="w").pack(fill="x")
        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, state='disabled', font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True)
        
        self.log_text.tag_config("INFO", foreground="black"); self.log_text.tag_config("SUCCESS", foreground="green")
        self.log_text.tag_config("ERROR", foreground="red"); self.log_text.tag_config("SETTING", foreground="blue")
        self.log_text.tag_config("WARN", foreground="#FF8C00"); self.log_text.tag_config("ENGINE", foreground="#FF00FF")
        self.log_text.tag_config("STATS", foreground="#008080")

        self.check_engines()

    # --- 辅助 ---
    def log(self, msg, level="INFO"):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, f"[{ts}] {msg}\n", level)
        self.log_text.see(tk.END); self.log_text.config(state='disabled')
        self.root.update()

    def clear_log(self):
        self.log_text.config(state='normal'); self.log_text.delete(1.0, tk.END); self.log_text.config(state='disabled')

    def get_column_letter(self, n):
        s=""; 
        while n>0: n,r=divmod(n-1,26); s=chr(65+r)+s
        return s

    def ask_open_folder(self, folder_path, message="操作完成！"):
        if messagebox.askyesno("完成", f"{message}\n\n是否立即打开文件夹查看？"):
            try: os.startfile(folder_path)
            except: pass

    # --- 引擎 ---
    def init_engine_panel(self):
        frame_eng = tk.LabelFrame(self.root, text="⚙️ 引擎配置", padx=10, pady=2, bg="#F0F8FF")
        frame_eng.pack(fill="x", padx=10, pady=(5, 0))
        self.lbl_status = tk.Label(frame_eng, text="检测中...", bg="#F0F8FF", font=("Arial", 9))
        self.lbl_status.pack(side="left", padx=10)
        tk.Label(frame_eng, text="| 引擎:", bg="#F0F8FF").pack(side="left", padx=5)
        rb_auto = tk.Radiobutton(frame_eng, text="🤖 自动", variable=self.engine_choice, value="auto", bg="#F0F8FF")
        rb_excel = tk.Radiobutton(frame_eng, text="🟢 Excel", variable=self.engine_choice, value="excel", bg="#F0F8FF")
        rb_wps = tk.Radiobutton(frame_eng, text="🔵 WPS", variable=self.engine_choice, value="wps", bg="#F0F8FF")
        rb_auto.pack(side="left"); rb_excel.pack(side="left"); rb_wps.pack(side="left")
        tk.Button(frame_eng, text="刷新", command=self.check_engines, width=8, bg="#E0E0E0").pack(side="right", padx=10)

    def check_engines(self):
        if not HAS_WIN32: self.lbl_status.config(text="❌ 未安装 pywin32", fg="red"); return
        self.has_excel=False; self.has_wps=False
        try: app=win32.Dispatch('Excel.Application'); app.Quit(); self.has_excel=True
        except: pass
        try: app=win32.Dispatch('Et.Application'); app.Quit(); self.has_wps=True
        except: 
            try: app=win32.Dispatch('Ket.Application'); app.Quit(); self.has_wps=True
            except: pass
        st = []
        if self.has_excel: st.append("Excel✅")
        if self.has_wps: st.append("WPS✅")
        if not st: st.append("无可用引擎❌")
        self.lbl_status.config(text=" ".join(st), fg="green" if (self.has_excel or self.has_wps) else "red")
        if self.has_excel: self.engine_choice.set("excel")
        elif self.has_wps: self.engine_choice.set("wps")

    def get_active_app_name(self):
        c = self.engine_choice.get()
        if c == "excel": return 'Excel.Application'
        if c == "wps": return 'Et.Application'
        if self.has_excel: return 'Excel.Application'
        if self.has_wps: return 'Et.Application'
        return None

    # --- 预览加载 (V40: 支持自定义行数) ---
    def refresh_preview(self):
        if self.current_preview_file: self.load_preview(self.current_preview_file)

    def load_preview(self, file_path):
        if not file_path or not os.path.exists(file_path): return
        self.current_preview_file = file_path
        
        try:
            limit = int(self.preview_limit.get())
            if limit < 1: limit = 50
        except: limit = 50

        try:
            self.lbl_preview_info.config(text=f"正在读取: {os.path.basename(file_path)} ...", fg="blue")
            self.root.update()
            
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            self.tree.delete(*self.tree.get_children())
            
            first_row = []
            for r in ws.iter_rows(min_row=1, max_row=1, values_only=True): first_row = list(r); break
            
            if not first_row: 
                self.lbl_preview_info.config(text=f"预览: [空文件]", fg="red"); wb.close(); return

            max_col = min(len(first_row), 30)
            cols = [str(i) for i in range(max_col + 1)]
            self.tree['columns'] = cols
            self.tree.column("0", width=40, anchor='center'); self.tree.heading("0", text="行号")
            for i in range(1, max_col + 1):
                c = self.get_column_letter(i)
                self.tree.column(str(i), width=100, anchor='w'); self.tree.heading(str(i), text=f"{c}")
            
            cnt = 0
            for i, r in enumerate(ws.iter_rows(min_row=1, max_row=limit, max_col=max_col, values_only=True)):
                vals = [i+1] + ["" if v is None else str(v) for v in r]
                tag = 'even' if i%2==0 else 'odd'
                self.tree.insert("", "end", values=vals, tags=(tag,))
                cnt += 1
            
            wb.close()
            self.lbl_preview_info.config(text=f"当前预览: {os.path.basename(file_path)} (显示前 {cnt} 行)", fg="#2E8B57")
        except Exception as e:
            self.log(f"预览失败: {e}", "ERROR"); self.lbl_preview_info.config(text="预览失败", fg="red")

    def on_preview_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region == "cell":
            item = self.tree.identify_row(event.y); vals = self.tree.item(item, 'values')
            if not vals: return
            row_idx = int(vals[0])
            tab = self.tab_control.index(self.tab_control.select())
            
            if tab == 0: 
                col = int(self.tree.identify_column(event.x).replace('#', '')) - 1
                if col>0: 
                    self.entry_split_col.delete(0, tk.END); self.entry_split_col.insert(0, str(col))
                    self.entry_start_row.delete(0, tk.END); self.entry_start_row.insert(0, str(row_idx))
                    self.log(f"拆分设定: 列{col}, 行{row_idx}", "SETTING")
            elif tab == 1:
                self.entry_merge_start_row.delete(0, tk.END); self.entry_merge_start_row.insert(0, str(row_idx))
                # [V39] 触发即时统计
                self.calculate_merge_stats(row_idx)

    # ================= V39 智能统计 =================
    def calculate_merge_stats(self, start_row):
        if not self.file_stats_cache:
            self.log("【提示】请先扫描文件夹，才能进行统计。", "WARN")
            return
        total_files = len(self.file_stats_cache)
        estimated_total_rows = 0
        for f_path, total_rows in self.file_stats_cache.items():
            valid = max(0, total_rows - start_row + 1)
            estimated_total_rows += valid
        self.log("-" * 40, "STATS")
        self.log(f"【即时统计】 数据开始行: {start_row}", "STATS")
        self.log(f"  - 参与文件: {total_files} 个", "STATS")
        self.log(f"  - 预计合并总行数: {estimated_total_rows} 行", "STATS")
        self.log("-" * 40, "STATS")

    def run_merge_report_thread(self):
        threading.Thread(target=self.generate_merge_report, daemon=True).start()

    def generate_merge_report(self):
        if not self.file_stats_cache: messagebox.showwarning("提示", "请先扫描文件夹"); return
        try: start_row = int(self.entry_merge_start_row.get())
        except: messagebox.showerror("错误", "起始行号无效"); return
        self.log("正在生成合并报告...", "INFO")
        report = []
        report.append("============ 📊 合并预估报告 ============")
        report.append(f"基准起始行: {start_row}")
        report.append(f"文件总数: {len(self.file_stats_cache)}")
        report.append("-" * 45)
        report.append(f"{'文件名':<30} | {'物理总行':<8} | {'预计贡献':<8}")
        report.append("-" * 45)
        total_valid = 0
        for f_path, total_rows in self.file_stats_cache.items():
            fname = os.path.basename(f_path)
            valid = max(0, total_rows - start_row + 1)
            total_valid += valid
            dname = (fname[:25] + '..') if len(fname) > 25 else fname
            report.append(f"{dname:<30} | {total_rows:<8} | {valid:<8}")
        report.append("-" * 45)
        report.append(f"【汇总】 预计合并后总行数: {total_valid}")
        top = tk.Toplevel(self.root); top.title("合并分析报告"); top.geometry("600x700")
        txt = scrolledtext.ScrolledText(top, font=("Consolas", 10))
        txt.pack(fill="both", expand=True, padx=10, pady=10)
        txt.insert(tk.END, "\n".join(report)); txt.config(state='disabled')
        self.log("合并报告已生成。", "SUCCESS")

    # ================= 拆分页面 =================
    def init_split_tab(self):
        frame = self.tab_split; frame_top = tk.Frame(frame); frame_top.pack(fill="x", padx=10, pady=10)
        tk.Label(frame_top, text="1. 选择总表:", font=("Microsoft YaHei", 10, "bold")).grid(row=0, column=0, sticky="w")
        self.entry_file_path = tk.Entry(frame_top, width=60); self.entry_file_path.grid(row=1, column=0, columnspan=2, sticky="w", pady=2)
        tk.Button(frame_top, text="浏览", command=self.select_file).grid(row=1, column=2, padx=5)
        f_btns = tk.Frame(frame_top); f_btns.grid(row=1, column=3, padx=5)
        tk.Button(f_btns, text="加载预览", command=lambda: self.load_preview(self.entry_file_path.get()), bg="#FFF176").pack(side="left", padx=2)
        tk.Button(f_btns, text="📊 分析报告", command=self.run_analysis_thread, bg="#B0E0E6").pack(side="left", padx=2)
        tk.Label(frame_top, text="2. 数据开始行:", font=("Microsoft YaHei", 10, "bold")).grid(row=2, column=0, sticky="w", pady=(5, 2))
        self.entry_start_row = tk.Entry(frame_top, width=15, bg="#F0F8FF"); self.entry_start_row.insert(0, "9"); self.entry_start_row.grid(row=3, column=0, sticky="w")
        tk.Label(frame_top, text="3. 拆分列号:", font=("Microsoft YaHei", 10, "bold")).grid(row=2, column=1, sticky="w", pady=(5, 2))
        self.entry_split_col = tk.Entry(frame_top, width=15, bg="#F0F8FF"); self.entry_split_col.insert(0, "3"); self.entry_split_col.grid(row=3, column=1, sticky="w")
        tk.Label(frame_top, text="👈 点击下方预览表自动填充", fg="blue").grid(row=2, column=2, rowspan=2, columnspan=2, sticky="w", padx=10)
        frame_middle = tk.Frame(frame); frame_middle.pack(fill="x", padx=10, pady=0)
        frame_mode = tk.LabelFrame(frame_middle, text="4. 拆分执行", padx=10, pady=5)
        frame_mode.pack(side="left", fill="both", expand=True, padx=(0, 5), pady=5)
        self.split_mode = tk.StringVar(value="perfect")
        tk.Radiobutton(frame_mode, text="极速", variable=self.split_mode, value="fast", fg="blue").pack(anchor="w")
        rb_perf = tk.Radiobutton(frame_mode, text="完美 (推荐)", variable=self.split_mode, value="perfect", fg="#8A2BE2")
        rb_perf.pack(anchor="w"); 
        if not HAS_WIN32: rb_perf.config(state="disabled")
        tk.Button(frame_mode, text="开始执行拆分", command=self.process_split, bg="#e1f5fe", height=1).pack(fill="x", pady=5)
        frame_tools = tk.LabelFrame(frame_middle, text="5. 实用工具箱", padx=10, pady=5, fg="#2E8B57")
        frame_tools.pack(side="left", fill="both", expand=True, padx=(5, 0), pady=5)
        tk.Button(frame_tools, text="🧹 强力清洗并另存", command=self.process_clean_save, bg="#98FB98", height=1).pack(anchor="center", fill="x", pady=10)

    # ================= 合并页面 =================
    def init_merge_tab(self):
        frame = self.tab_merge
        frame_top = tk.Frame(frame); frame_top.pack(fill="x", padx=20, pady=10)
        tk.Label(frame_top, text="1. 选择回收文件夹 (自动扫描):", font=("Microsoft YaHei", 10, "bold")).grid(row=0, column=0, sticky="w")
        self.entry_folder_path = tk.Entry(frame_top, width=50); self.entry_folder_path.grid(row=1, column=0, columnspan=2, sticky="w", pady=2)
        tk.Button(frame_top, text="浏览...", command=self.select_folder_and_scan).grid(row=1, column=2, padx=5)
        
        # 统计按钮区
        frame_btns = tk.Frame(frame_top); frame_btns.grid(row=2, column=0, columnspan=4, sticky="w", pady=5)
        tk.Label(frame_btns, text="2. 数据开始行:", font=("Microsoft YaHei", 10, "bold")).pack(side="left")
        self.entry_merge_start_row = tk.Entry(frame_btns, width=10, bg="#F0F8FF"); self.entry_merge_start_row.insert(0, "9"); self.entry_merge_start_row.pack(side="left", padx=5)
        tk.Button(frame_btns, text="📉 刷新统计", command=lambda: self.calculate_merge_stats(int(self.entry_merge_start_row.get() or 9)), bg="#E0FFFF").pack(side="left", padx=5)
        tk.Button(frame_btns, text="📊 合并报告", command=self.run_merge_report_thread, bg="#B0E0E6").pack(side="left", padx=5)

        self.lbl_template = tk.Label(frame_top, text="当前模板: [未选择] (默认首个)", fg="gray"); self.lbl_template.grid(row=3, column=1, columnspan=2, sticky="w", padx=10)
        tk.Button(frame_top, text="开始合并", command=self.process_merge, bg="#e1f5fe", height=2, width=20).grid(row=2, column=3, rowspan=2, padx=10)

        frame_list = tk.LabelFrame(frame, text="📄 待合并文件列表 (点击预览)", padx=10, pady=5)
        frame_list.pack(fill="both", expand=True, padx=20, pady=5)

        self.file_tree = ttk.Treeview(frame_list, show='headings', height=6)
        self.file_tree['columns'] = ("idx", "filename", "total_rows", "valid_rows", "empty_rows", "cols", "size")
        self.file_tree.tag_configure('odd', background='white'); self.file_tree.tag_configure('even', background='#f9f9f9')
        self.file_tree.column("idx", width=40, anchor='center'); self.file_tree.heading("idx", text="序号")
        self.file_tree.column("filename", width=200, anchor='w'); self.file_tree.heading("filename", text="文件名")
        self.file_tree.column("total_rows", width=70); self.file_tree.heading("total_rows", text="总行数")
        self.file_tree.column("valid_rows", width=70); self.file_tree.heading("valid_rows", text="有效行")
        self.file_tree.column("empty_rows", width=60); self.file_tree.heading("empty_rows", text="空行")
        self.file_tree.column("cols", width=50); self.file_tree.heading("cols", text="列数")
        self.file_tree.column("size", width=80); self.file_tree.heading("size", text="大小")

        vsb_f = ttk.Scrollbar(frame_list, orient="vertical", command=self.file_tree.yview); self.file_tree.configure(yscrollcommand=vsb_f.set)
        self.file_tree.pack(side="left", fill="both", expand=True); vsb_f.pack(side="right", fill="y")
        self.file_tree.bind('<ButtonRelease-1>', self.on_file_list_click)
        self.file_tree.bind('<Double-1>', self.on_file_list_double_click)

    def select_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if f: self.entry_file_path.delete(0, tk.END); self.entry_file_path.insert(0, f); self.load_preview(f)
    def select_folder_and_scan(self):
        d = filedialog.askdirectory()
        if d: self.entry_folder_path.delete(0, tk.END); self.entry_folder_path.insert(0, d); threading.Thread(target=self.scan_merge_folder, daemon=True).start()

    # --- V39 全景扫描逻辑 ---
    def scan_merge_folder(self):
        folder = self.entry_folder_path.get()
        if not folder: return
        self.clear_log(); self.log(f"正在全景扫描: {folder}", "INFO")
        for item in self.file_tree.get_children(): self.file_tree.delete(item)
        self.merge_files_cache = []; self.file_stats_cache = {}; self.current_template = None; self.lbl_template.config(text="模板: [未选择]", fg="gray")
        files = glob.glob(os.path.join(folder, "*.xlsx"))
        files = [f for f in files if "汇总" not in os.path.basename(f) and not os.path.basename(f).startswith("~$")]
        if not files: self.log("未找到 .xlsx 文件", "WARN"); return
        self.log(f"发现 {len(files)} 个文件，开始深度分析...", "INFO")
        for idx, f in enumerate(files):
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
                total = 0; valid = 0; max_c = 0
                for i, r in enumerate(ws.iter_rows(values_only=True)):
                    if i > 30000: total = "30000+"; break
                    total = i + 1
                    if len(r) > max_c: max_c = len(r)
                    if any(c is not None and str(c).strip() != "" for c in r): valid += 1
                wb.close()
                empty = 0
                if isinstance(total, int): empty = total - valid
                self.file_stats_cache[f] = total if isinstance(total, int) else 30000 # 缓存物理总行
                sz = f"{round(os.path.getsize(f)/1024, 1)} KB"
                tag = 'even' if idx%2==0 else 'odd'
                self.file_tree.insert("", "end", values=(idx+1, os.path.basename(f), total, valid, empty, max_c, sz), tags=(tag,))
                self.merge_files_cache.append(f)
                if idx%5==0: self.root.update()
            except Exception as e: self.log(f"扫描失败: {os.path.basename(f)}", "ERROR")
        self.log("全景扫描完成。", "SUCCESS")
        if self.merge_files_cache: self.set_template(0)

    def on_file_list_click(self, event):
        row_id = self.file_tree.identify_row(event.y)
        if not row_id: return
        self.file_tree.selection_set(row_id)
        idx = int(self.file_tree.item(row_id, 'values')[0]) - 1
        self.set_template(idx)
    def on_file_list_double_click(self, event):
        row_id = self.file_tree.identify_row(event.y)
        if not row_id: return
        idx = int(self.file_tree.item(row_id, 'values')[0]) - 1
        if 0 <= idx < len(self.merge_files_cache):
            try: os.startfile(self.merge_files_cache[idx])
            except: pass
    def set_template(self, idx):
        if 0 <= idx < len(self.merge_files_cache):
            path = self.merge_files_cache[idx]
            self.current_template = path
            self.lbl_template.config(text=f"当前模板: {os.path.basename(path)}", fg="#8A2BE2")
            self.load_preview(path)

    # --- 核心：修复与清理 ---
    def sanitize_file(self, file_path):
        if not HAS_WIN32: return file_path
        abs_path = os.path.abspath(file_path)
        temp_dir = os.path.join(os.path.dirname(abs_path), "_temp_repair")
        if not os.path.exists(temp_dir): os.makedirs(temp_dir)
        temp_path = os.path.join(temp_dir, os.path.basename(file_path).split('.')[0] + "_shadow.xlsx")
        prog = self.get_active_app_name()
        if not prog: return file_path
        self.log(f"影子缓存 ({prog})...", "INFO")
        app = None
        try:
            try: app = win32.Dispatch(prog)
            except: app = win32.Dispatch('Ket.Application')
            app.Visible = False; app.DisplayAlerts = False
            wb = app.Workbooks.Open(abs_path); wb.SaveAs(temp_path, FileFormat=51); wb.Close(); app.Quit()
            return temp_path
        except: 
            if app: 
                try: app.Quit() 
                except: pass
            return file_path

    def process_clean_save(self):
        f = self.entry_file_path.get()
        if not f: return
        self.clear_log(); self.log("开始强力清洗...", "INFO")
        if not HAS_WIN32: messagebox.showerror("错", "需Excel/WPS"); return
        prog = self.get_active_app_name(); 
        if not prog: return
        threading.Thread(target=self.run_native_clean, args=(f, prog), daemon=True).start()

    def run_native_clean(self, file_path, prog_id):
        app = None
        try:
            try: app = win32.Dispatch(prog_id)
            except: app = win32.Dispatch('Ket.Application')
            app.Visible = False; app.DisplayAlerts = False
            dir_name = os.path.dirname(file_path); base_name = os.path.basename(file_path)
            ts = datetime.datetime.now().strftime("%H%M%S")
            nm = f"{os.path.splitext(base_name)[0]}_清洗_{ts}.xlsx"
            wb = app.Workbooks.Open(os.path.abspath(file_path)); ws = wb.ActiveSheet
            mr = ws.UsedRange.Rows.Count + ws.UsedRange.Row - 1
            mc = ws.UsedRange.Columns.Count + ws.UsedRange.Column - 1
            self.log("扫描并删除空行/列...", "INFO")
            for r in range(mr, 0, -1):
                if app.WorksheetFunction.CountA(ws.Rows(r)) == 0: ws.Rows(r).Delete()
            for c in range(mc, 0, -1):
                if app.WorksheetFunction.CountA(ws.Columns(c)) == 0: ws.Columns(c).Delete()
            wb.SaveAs(os.path.join(dir_name, nm), FileFormat=51); wb.Close(); app.Quit()
            self.log(f"完成: {nm}", "SUCCESS")
            self.ask_open_folder(dir_name, f"清洗完成: {nm}")
        except Exception as e:
            self.log(f"清洗失败: {e}", "ERROR")
            if app: 
                try: app.Quit() 
                except: pass

    # --- 拆分 ---
    def run_analysis_thread(self): threading.Thread(target=self.generate_analysis_report, daemon=True).start()
    def generate_analysis_report(self):
        f = self.entry_file_path.get()
        if not f: return
        try: start_row=int(self.entry_start_row.get()); col_idx=int(self.entry_split_col.get())
        except: return
        self.log("分析中...", "INFO")
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
            total, data_cnt, vals = 0, 0, []
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i>20000: total="20000+"; break
                total=i+1
                if i+1>=start_row:
                    v = r[col_idx-1] if col_idx-1<len(r) else None
                    if v: vals.append(str(v).strip())
            wb.close()
            counter = Counter(vals)
            rep = [f"文件: {os.path.basename(f)}", f"扫描: {total}", f"有效: {data_cnt}", "-"*30, "【分类预览】"]
            for k, v in counter.most_common(): rep.append(f"{k} : {v}")
            top = tk.Toplevel(self.root); txt = scrolledtext.ScrolledText(top); txt.pack(fill="both")
            txt.insert(tk.END, "\n".join(rep))
        except Exception as e: self.log(f"错: {e}", "ERROR")

    def process_split(self):
        self.clear_log()
        f = self.entry_file_path.get()
        if not f: return
        try: start_row=int(self.entry_start_row.get()); col_idx=int(self.entry_split_col.get())
        except: return
        mode = self.split_mode.get()
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = os.path.join(os.path.dirname(f), f"拆分结果_{ts}")
        if not os.path.exists(out_dir): os.makedirs(out_dir)
        try:
            cnt = 0
            if mode == "fast": cnt = self.run_fast_split(f, start_row, col_idx, out_dir, ts)
            else: cnt = self.run_perfect_split(f, start_row, col_idx, out_dir, ts)
            self.ask_open_folder(out_dir, f"拆分完成！生成 {cnt} 个文件。")
        except Exception as e: self.log(f"错: {e}", "ERROR")

    def run_fast_split(self, f, start, col, out, ts):
        self.log("极速模式...", "INFO")
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
        h, d = [], []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i+1 < start: h.append(r)
            else: d.append(r)
        wb.close()
        m = {}
        for r in d:
            v = r[col-1] if col-1 < len(r) else None
            if v: 
                if v not in m: m[v] = []
                m[v].append(r)
        cnt = 0
        for k, v in m.items():
            n = str(k).replace('/', '_').strip()
            nb = openpyxl.Workbook(); ns = nb.active
            for r in h: ns.append(r)
            for r in v: ns.append(r)
            ns.sheet_format.defaultRowHeight = 25
            nb.save(os.path.join(out, f"{n}_极速_{ts}.xlsx")); self.log(f"生成: {n}", "SUCCESS"); cnt+=1
        return cnt

    def run_perfect_split(self, original_file, start_row, col_idx, output_dir, timestamp):
        prog_id = self.get_active_app_name()
        if not prog_id: self.log("无引擎", "ERROR"); return
        self.log(f">>> 完美模式 ({prog_id})...", "ENGINE")
        shadow_file = self.sanitize_file(original_file)
        temp_made = (shadow_file != original_file)
        wb_scan = openpyxl.load_workbook(shadow_file, read_only=True, data_only=True); ws_scan = wb_scan.active
        real_max_row = 0; row_data_map = {}
        for i, r in enumerate(ws_scan.iter_rows(values_only=True)):
            row_num = i+1
            if row_num >= start_row:
                val = r[col_idx-1] if col_idx-1<len(r) else None
                if val: real_max_row = row_num; row_data_map[row_num] = str(val).strip()
        wb_scan.close()
        targets = set(row_data_map.values())
        self.log(f"有效数据截止: {real_max_row} 行", "INFO")
        app = None; count = 0
        try:
            try: app = win32.Dispatch(prog_id)
            except: app = win32.Dispatch('Ket.Application')
            app.Visible = False; app.DisplayAlerts = False
            _, ext = os.path.splitext(original_file)
            for idx, target_val in enumerate(targets):
                safe_name = str(target_val).replace('/', '_').strip()
                t_file = f"{safe_name}{ext}"
                t_path = os.path.join(output_dir, t_file)
                self.log(f"[{idx+1}/{len(targets)}] {t_file}", "INFO")
                shutil.copy2(original_file, t_path)
                wb = app.Workbooks.Open(os.path.abspath(t_path)); ws = wb.ActiveSheet
                if real_max_row < 1048576:
                    try: ws.Range(f"A{real_max_row+1}:A1048576").EntireRow.Delete()
                    except: pass
                app.ScreenUpdating = False
                del_rng = None; bat = 0
                for r in range(real_max_row, start_row-1, -1):
                    owner = row_data_map.get(r)
                    should_del = False
                    if owner and owner != target_val: should_del = True
                    elif not owner: should_del = True 
                    if should_del:
                        if not del_rng: del_rng = ws.Rows(r)
                        else: del_rng = app.Union(del_rng, ws.Rows(r))
                        bat += 1
                    if bat >= 50: del_rng.Delete(); del_rng = None; bat = 0
                if del_rng: del_rng.Delete()
                app.ScreenUpdating = True
                wb.Save(); wb.Close(); count += 1
            app.Quit()
        except Exception as e:
            self.log(f"引擎错: {e}", "ERROR"); 
            if app: 
                try: app.Quit() 
                except: pass
        if temp_made: 
            try: shutil.rmtree(os.path.dirname(shadow_file)) 
            except: pass
        return count

    # --- 合并 ---
    def process_merge(self):
        folder = self.entry_folder_path.get()
        if not folder: return
        if not self.merge_files_cache: self.scan_merge_folder()
        files = self.merge_files_cache
        if not files: return
        try: start_row = int(self.entry_merge_start_row.get())
        except: return
        templ = self.current_template if self.current_template else files[0]
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log(f"合并中... 模板: {os.path.basename(templ)}", "INFO")
        try:
            save_name = f"合并汇总表_{ts}.xlsx"
            save_path = os.path.join(folder, save_name)
            if HAS_WIN32:
                # 为了防止模板带脏数据，可以先清洗再copy
                # 但为了速度，这里直接copy，假设模板是好的
                shutil.copy2(templ, save_path)
            else: shutil.copy2(templ, save_path)
            main_wb = openpyxl.load_workbook(save_path); main_ws = main_wb.active
            if main_ws.max_row >= start_row: main_ws.delete_rows(start_row, amount=main_ws.max_row - start_row + 1)
            cnt = 0
            for idx, f in enumerate(files):
                self.log(f"[{idx+1}/{len(files)}] 读取: {os.path.basename(f)}", "INFO")
                wb_src = openpyxl.load_workbook(f, read_only=True, data_only=True); ws_src = wb_src.active
                for i, r in enumerate(ws_src.iter_rows(values_only=True)):
                    if i+1 >= start_row and any(c is not None and str(c).strip()!="" for c in r):
                        main_ws.append(r); cnt+=1
                wb_src.close()
            main_ws.sheet_format.defaultRowHeight = 25
            main_wb.save(save_path)
            self.ask_open_folder(os.path.dirname(save_path), f"合并完成！共 {cnt} 行。")
        except Exception as e: self.log(f"合并错: {e}", "ERROR")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelToolApp(root)
    root.mainloop()